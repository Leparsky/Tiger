#https://habr.com/post/322608/
#https://habr.com/post/250921/
from selenium import webdriver
import selenium.common.exceptions
import time
running = True
from selenium.webdriver.chrome.options import DesiredCapabilities
from selenium.webdriver.common.proxy import Proxy, ProxyType

from openpyxl import Workbook

BASE_URL = 'https://ajento.ru'     #адрес сайта для парсинга
ALL_PROXIES =[]
driver = None



#https://gist.github.com/tushortz/cba8b25f9d80f584f807b65890f37be5
def get_proxies():
    co = webdriver.ChromeOptions()
    co.add_argument("log-level=3")
    co.add_argument("--headless")
    driver = webdriver.Chrome(chrome_options=co)
    driver.get("https://free-proxy-list.net/")

    PROXIES = []
    #PROXIES.append("91.137.129.85:36784")
    #PROXIES.append("60.52.120.97:8080")
    #PROXIES.append("77.45.14.72:43264")
    PROXIES.append("187.188.175.205:32978")
    PROXIES.append("118.175.176.135:45028")
    PROXIES.append("134.249.156.3:35775")
    PROXIES.append("141.193.188.97:30416")
    PROXIES.append("93.178.206.36:59339")
    PROXIES.append("170.245.224.148:57276")
    PROXIES.append("96.21.217.66:42873")
    PROXIES.append("194.9.26.237:39729")
    PROXIES.append("94.21.151.254:42378")
    PROXIES.append("195.230.10.212:36921")
    PROXIES.append("81.196.94.238:51560")
    PROXIES.append("110.235.250.75:31364")
    PROXIES.append("178.216.190.53:35863")
    PROXIES.append("95.31.41.216:55546")
    PROXIES.append("213.21.174.202:61574")
    PROXIES.append("188.2.49.63:32573")
    PROXIES.append("1.2.169.79:47096")
    PROXIES.append("103.9.190.237:35381")
    PROXIES.append("108.29.37.161:59481")
    PROXIES.append("95.167.65.6:54605")
    PROXIES.append("87.235.186.171:46622")
    PROXIES.append("159.253.81.90:39530")
    PROXIES.append("116.254.112.180:52566")
    PROXIES.append("103.247.121.114:37278")
    PROXIES.append("115.85.67.130:45764")
    PROXIES.append("212.115.232.201:39958")
    PROXIES.append("188.94.140.2:59011")
    PROXIES.append("193.33.8.18:52139")
    PROXIES.append("91.211.190.64:30980")
    PROXIES.append("178.165.120.54:39753")
    PROXIES.append("181.129.49.26:53750")
    PROXIES.append("200.44.166.30:57972")
    PROXIES.append("94.158.151.181:35675")
    PROXIES.append("167.114.79.139:59819")
    PROXIES.append("103.255.240.251:52935")
    PROXIES.append("110.74.195.62:55585")
    PROXIES.append("101.109.255.59:53502")
    PROXIES.append("77.242.24.91:59448")
    PROXIES.append("189.142.158.26:33485")
    PROXIES.append("36.89.74.130:45042")
    PROXIES.append("77.95.193.131:33284")
    PROXIES.append("177.193.221.111:44147")
    PROXIES.append("87.254.138.159:43710")
    PROXIES.append("178.165.66.92:60159")

    proxies = driver.find_elements_by_css_selector("tr[role='row']")
    for p in proxies:
        result = p.text.split(" ")

        if result[-1] == "yes":
            PROXIES.append(result[0]+":"+result[1])

    driver.close()
    #PROXIES=PROXIES[5:-3]
    return PROXIES





def proxy_driver(PROXIES):
    co = webdriver.ChromeOptions()
    co.add_argument("log-level=3")
    #co.add_argument("--headless")
    prox = Proxy()

    if PROXIES:
        pxy = PROXIES[-1]
    else:
        print("--- Proxies used up (%s)" % len(PROXIES))
        PROXIES = get_proxies()

    prox.proxy_type = ProxyType.MANUAL
    prox.http_proxy = pxy
    prox.socks_proxy = pxy
    prox.ssl_proxy = pxy

    capabilities = webdriver.DesiredCapabilities.CHROME
    prox.add_to_capabilities(capabilities)

    driver = webdriver.Chrome(chrome_options=co, desired_capabilities=capabilities)

    return driver
def get_page(url):
    global ALL_PROXIES
    global driver
    running = True
    while running:
        try:
            driver.get(url)

            assert "Телефоны"  in driver.page_source
            running = False
            #running = True
        except:
            new = ALL_PROXIES.pop()
            # reassign driver if fail to switch proxy
            driver.close()
            driver = proxy_driver(ALL_PROXIES)
            print("--- Switched proxy to: %s" % new)
            time.sleep(1)

def main():
    global ALL_PROXIES
    global driver
    ALL_PROXIES = get_proxies()
    driver = proxy_driver(ALL_PROXIES)
    driver.maximize_window()
    lstCatPages = []
    # страницы каталога
    get_page(BASE_URL)

    for element in driver.find_elements_by_xpath("//a[@class='cs-sub-menu__link']"):
        lstCatPages.append(element.get_attribute("href")) if element.get_attribute("href") not in lstCatPages else None
        # добавляем в список вложеные подменю проверяем задвоение
    for element in driver.find_elements_by_xpath("//a[@class='cs-menu__link']"):
        lstCatPages.append(element.get_attribute("href")) if element.get_attribute("href") not in lstCatPages else None
        # добавляем в список верхние меню, проверяем задвоение
    print("количество страниц каталога :" +str(len(lstCatPages)))
#

    lstGoodPages = []
    # страницы каталога товаров

    for page in lstCatPages:
        get_page(page)
        for element in driver.find_elements_by_xpath("//a[@class='cs-goods-title']"):
            lstGoodPages.append(element.get_attribute("href")) if element.get_attribute("href") not in lstGoodPages else None
            #добавляем в список товаров ссылку,если такой ссылки нет, проверяем задвоение
        for element in driver.find_elements_by_xpath("//a[@class=' b-pager__link ']"):
            lstCatPages.append(element.get_attribute("href")) if element.get_attribute("href") not in lstCatPages else None
            # добавляем в список пейджеры, проверяем задвоение
        print("количество страниц каталога :" + str(len(lstCatPages)))
        print("количество страниц товаров :" + str(len(lstGoodPages)))


    #lstGoodPages.append("https://ajento.ru/p346901069-black-stone-modnaya.html")

    wb = Workbook()
    ws = wb.active
    pagesproxycount = 30
    ws.cell(column=1+0, row=1, value="ID")
    ws.cell(column=1+1, row=1, value="Название")
    ws.cell(column=1+2, row=1, value="Оригинальное")
    ws.cell(column=1+3, row=1, value="название")
    ws.cell(column=1+4, row=1, value="Цена")
    ws.cell(column=1+5, row=1, value="Количество")
    ws.cell(column=1+6, row=1, value="Размер")
    ws.cell(column=1+7, row=1, value="Артикул")
    ws.cell(column=1+8, row=1, value="Цвет")
    ws.cell(column=1+9, row=1, value="Единица измерения")
    ws.cell(column=1+10, row=1, value="Описание")
    ws.cell(column=1+11, row=1, value="Фотография")
    ws.cell(column=1+12, row=1, value="Альбом")
    ws.cell(column=1+13, row=1, value="Позиция")
    ws.cell(column=1+14, row=1, value="Отображать комментарий")
    ws.cell(column=1+15, row=1, value="Отображать в каталоге")
    ws.cell(column=1+16, row=1, value="Включить ряды")
    ws.cell(column=1+17, row=1, value="Ссылка на источник")
    ws.cell(column=1+18, row=1, value="Обновить фото")


    i=1
    for page in lstGoodPages:
        url = ""  # url
        cost = ""  # Цена
        size = ""  # Размер
        art = ""  # Артикул
        color = ""  # Цвет
        edizm = ""  # Единица    измерения
        descr = ""  # // Описание
        album = ""  # // Альбом
        position = ""  # Позиция
        picture = ""  # // Ссылка на картинку
        name = ""  # Название
        get_page(page)
        i+=1
        print(page)
        url = page

        element = driver.find_element_by_xpath("//p[@class='b-product-cost__price']")
        cost=(element.text)
        size = ""
        for element in driver.find_elements_by_xpath("//li[ @class='b-custom-drop-down__list-item']"):
            size+=element.get_attribute("innerHTML")+','

        #size1 = ""
        for element in driver.find_elements_by_xpath("//span[@class='b-product-mods__button-text']")[1:]:
            size += element.text+','
        size=size[:-1]
        element = driver.find_element_by_xpath("//h1[@class='cs-title cs-title_type_product cs-online-edit']/span")
        name = (element.text)

        element = driver.find_element_by_xpath("//li[@class='b-product-data__item b-product-data__item_type_sku']/span")
        art = (element.text)
        try:
            element = driver.find_element_by_xpath("//li[@class='b-product-data__item b-product-data__item_type_available']")
            descr = (element.text)+' , '
            element = driver.find_element_by_xpath("//table[@class='b-product-info']")
            descr += element.get_attribute('textContent')
        except selenium.common.exceptions.NoSuchElementException:
            None
        try:
                element = driver.find_element_by_xpath("//li[@class='b-product-data__item b-product-data__item_type_selling']")
                descr += (element.text)
        except selenium.common.exceptions.NoSuchElementException:
                None
        try:
            element = driver.find_element_by_xpath("//p[@class='b-product-cost__min-order']")
            descr += (element.text)
        except selenium.common.exceptions.NoSuchElementException:
            None
        try:
            element = driver.find_element_by_xpath("//img[@class='cs-product-image__img']")
            picture += element.get_attribute("src")
        except selenium.common.exceptions.NoSuchElementException:
            None
        ws.cell(column=1+1, row=i, value=name)
        ws.cell(column=1+4, row=i, value=cost)
        ws.cell(column=1+6, row=i, value=size)
        ws.cell(column=1+7, row=i, value=art)
        ws.cell(column=1+8, row=i, value=color)
        ws.cell(column=1+9, row=i, value=edizm)
        ws.cell(column=1+10,row=i, value=descr)
        ws.cell(column=1+12,row=i, value=album)
        ws.cell(column=1+13,row=i, value=i - 1)
        ws.cell(column=1+17,row=i, value=picture)
        ws.cell(column=1+18,row=i, value=url)
    wb.save(filename="c:\\tmp\\{}ajento.xlsx".format(time.time()))


main()
#запуск приложения
